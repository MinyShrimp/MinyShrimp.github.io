
import openpyxl, json

if __name__ == '__main__':
    wb = openpyxl.load_workbook('./data.xlsx')
    ws = wb['1340 무기']

    _result = {}
    for _row in range(2, 21):
        _result[ ws['A{}'.format(_row)].value ] = {
            'success':       ws['B{}'.format(_row)].value,
            'fail_up':       ws['C{}'.format(_row)].value,
            'suho':          ws['D{}'.format(_row)].value,
            'wemeng':        ws['E{}'.format(_row)].value,
            'oreha':         ws['F{}'.format(_row)].value,
            'mengpa':        ws['G{}'.format(_row)].value,
            'gold':          ws['H{}'.format(_row)].value,
            'sun1':          ws['I{}'.format(_row)].value,
            'sun1_count':    ws['J{}'.format(_row)].value,
            'sun2':          ws['K{}'.format(_row)].value,
            'sun2_count':    ws['L{}'.format(_row)].value,
            'sun3':          ws['M{}'.format(_row)].value,
            'sun3_count':    ws['N{}'.format(_row)].value,
        }
    with open("./weapon.json", "w") as json_file:
        json.dump(_result, json_file, indent=4)
    
    ws = wb['1340 방어구']
    _result = {}
    for _row in range(2, 21):
        _result[ ws['A{}'.format(_row)].value ] = {
            'success':       ws['B{}'.format(_row)].value,
            'fail_up':       ws['C{}'.format(_row)].value,
            'suho':          ws['D{}'.format(_row)].value,
            'wemeng':        ws['E{}'.format(_row)].value,
            'oreha':         ws['F{}'.format(_row)].value,
            'mengpa':        ws['G{}'.format(_row)].value,
            'gold':          ws['H{}'.format(_row)].value,
            'sun1':          ws['I{}'.format(_row)].value,
            'sun1_count':    ws['J{}'.format(_row)].value,
            'sun2':          ws['K{}'.format(_row)].value,
            'sun2_count':    ws['L{}'.format(_row)].value,
            'sun3':          ws['M{}'.format(_row)].value,
            'sun3_count':    ws['N{}'.format(_row)].value,
        }
    with open("./equip.json", "w") as json_file:
        json.dump(_result, json_file, indent=4)